"""
UP Invest - Government Orders (Revision Petitions) Downloader & Extractor
Downloads all judgement PDFs and creates an Excel sheet with structured data.
PDFs are scanned images — Claude Vision is used for OCR + field extraction.
"""

import os
import sys
import time
import json
import warnings
import pathlib
import re
import base64
import getpass

import requests
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup
import anthropic

warnings.filterwarnings("ignore")  # suppress SSL warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL = "https://invest.up.gov.in"
LISTING_URL = "https://invest.up.gov.in/gos/?pagenum={page}"
DOWNLOAD_DIR = pathlib.Path("UP_Judgements")
EXCEL_FILE = pathlib.Path("UP_Judgements_Index.xlsx")
PROGRESS_FILE = pathlib.Path("progress.json")
MAX_PAGES = 50  # safety cap; actual count detected dynamically
MAX_PDF_PAGES = 12  # max pages to send to Claude per document

SESSION = requests.Session()
SESSION.verify = False
SESSION.headers["User-Agent"] = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
)


# ── helpers ──────────────────────────────────────────────────────────────────

def get_total_pages() -> int:
    r = SESSION.get(LISTING_URL.format(page=1), timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    nums = [
        int(a.get_text(strip=True))
        for a in soup.find_all("a", class_="page-numbers")
        if a.get_text(strip=True).isdigit()
    ]
    return max(nums) if nums else 1


def scrape_page(page: int) -> list[dict]:
    r = SESSION.get(LISTING_URL.format(page=page), timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("table", id="ContentPlaceHolder_Body_gvgovernOrder")
    if not table:
        return []
    rows = []
    for tr in table.find_all("tr")[1:]:  # skip header row
        cells = tr.find_all("td")
        if len(cells) < 3:
            continue
        a = cells[1].find("a")
        if not a:
            continue
        rows.append(
            {
                "title": a.get_text(strip=True),
                "pdf_url": a["href"].strip(),
                "date_listed": cells[2].get_text(strip=True),
            }
        )
    return rows


def download_pdf(pdf_url: str, dest: pathlib.Path) -> bool:
    if dest.exists() and dest.stat().st_size > 0:
        return True  # already downloaded
    try:
        r = SESSION.get(pdf_url, timeout=90, stream=True)
        r.raise_for_status()
        dest.write_bytes(r.content)
        return True
    except Exception as e:
        print(f"  ✗ Download failed: {e}")
        return False


def pdf_pages_to_base64_images(pdf_path: pathlib.Path, max_pages: int = MAX_PDF_PAGES) -> list[str]:
    """Convert PDF pages to base64-encoded PNG images using PyMuPDF."""
    images = []
    try:
        doc = fitz.open(str(pdf_path))
        n = min(len(doc), max_pages)
        for page_num in range(n):
            page = doc[page_num]
            # 200 DPI is good enough for OCR and keeps image size reasonable
            mat = fitz.Matrix(200 / 72, 200 / 72)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            png_bytes = pix.tobytes("png")
            images.append(base64.standard_b64encode(png_bytes).decode("utf-8"))
        doc.close()
    except Exception as e:
        print(f"  ✗ Image conversion error: {e}")
    return images


def extract_fields_with_claude_vision(
    client: anthropic.Anthropic,
    images_b64: list[str],
    title: str,
    date_listed: str,
) -> dict:
    """Send page images to Claude and extract structured fields."""

    content = []

    # Add all page images
    for img_b64 in images_b64:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": img_b64,
            },
        })

    content.append({
        "type": "text",
        "text": f"""You are a legal analyst. The images above are pages from a government order (revision petition decision) from the Uttar Pradesh government under the UP Urban Planning and Development Act.

Read all pages carefully and extract the following fields. Return ONLY a valid JSON object with these exact keys:

- "case_name": Full name of the case (e.g. "XYZ Pvt Ltd vs NOIDA Authority"). Include both parties.
- "date_of_judgement": Date the order was passed (DD-MM-YYYY format). Look for the date near the signature or at the top.
- "author": Full name and designation of the officer/authority who signed/passed the order (e.g. "Shri Amit Kumar, Principal Secretary, Housing & Urban Planning").
- "lawyers": Comma-separated list of advocates/lawyers who appeared for both sides. Include party names too if mentioned (e.g. "Adv. Ramesh Gupta for petitioner, Adv. Sunil Sharma for respondent"). If not mentioned, write "Not mentioned".
- "brief_facts": 3-5 clear sentences summarising: what the original dispute was about, what the petitioner sought, and what happened at the authority level before the revision.
- "relief_granted": The final decision — what was ordered, whether the petition was allowed/dismissed, and any specific directions given.

Website title: {title}
Listed date: {date_listed}

Return only the JSON object, no markdown fences, no explanation.""",
    })

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1200,
            messages=[{"role": "user", "content": content}],
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```[a-z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        return json.loads(raw)
    except json.JSONDecodeError:
        # Try to salvage partial JSON
        return {
            "case_name": title,
            "date_of_judgement": date_listed,
            "author": "Parse error",
            "lawyers": "Parse error",
            "brief_facts": f"Could not parse Claude response.",
            "relief_granted": "Parse error",
        }
    except Exception as e:
        return {
            "case_name": title,
            "date_of_judgement": date_listed,
            "author": "Error",
            "lawyers": "Error",
            "brief_facts": f"Extraction error: {e}",
            "relief_granted": "Error",
        }


# ── Excel output ──────────────────────────────────────────────────────────────

COLUMNS = [
    ("Sr No.", 8),
    ("Case Name", 45),
    ("Date of Judgement", 18),
    ("Author of Order", 35),
    ("Lawyers Who Appeared", 45),
    ("Brief Facts", 70),
    ("Relief Granted", 55),
    ("PDF File", 42),
    ("Source URL", 55),
]


def create_workbook() -> tuple[openpyxl.Workbook, object]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UP Revision Petitions"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = col_width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    return wb, ws


def append_row(ws, sr: int, fields: dict, pdf_filename: str, pdf_url: str):
    row = [
        sr,
        fields.get("case_name", ""),
        fields.get("date_of_judgement", ""),
        fields.get("author", ""),
        fields.get("lawyers", ""),
        fields.get("brief_facts", ""),
        fields.get("relief_granted", ""),
        pdf_filename,
        pdf_url,
    ]
    ws.append(row)
    r = ws.max_row
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r % 2 == 0:
            cell.fill = alt_fill
    ws.row_dimensions[r].height = 90


# ── progress tracking ─────────────────────────────────────────────────────────

def load_progress() -> dict:
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text())
    return {"done_urls": [], "last_sr": 0}


def save_progress(progress: dict):
    PROGRESS_FILE.write_text(json.dumps(progress, indent=2))


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  UP Government Orders - Judgement Downloader & Analyser")
    print("  (Uses Claude Vision for OCR on scanned PDFs)")
    print("=" * 65)

    # API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        api_key = getpass.getpass("Enter your Anthropic API key: ").strip()
    if not api_key:
        print("ERROR: API key is required.")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    DOWNLOAD_DIR.mkdir(exist_ok=True)

    # Load or create workbook
    progress = load_progress()
    done_urls: set = set(progress["done_urls"])
    sr = progress["last_sr"]

    if EXCEL_FILE.exists() and done_urls:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        print(f"Resuming — {len(done_urls)} already processed.\n")
    else:
        wb, ws = create_workbook()
        print("Starting fresh.\n")

    # Discover pages
    print("Detecting total pages...", end=" ", flush=True)
    try:
        total_pages = min(get_total_pages(), MAX_PAGES)
    except Exception as e:
        print(f"failed ({e}), defaulting to {MAX_PAGES}")
        total_pages = MAX_PAGES
    print(f"{total_pages} pages found.")

    # Collect all entries
    print("Collecting listing entries...")
    all_entries = []
    for page in range(1, total_pages + 1):
        try:
            entries = scrape_page(page)
            all_entries.extend(entries)
            print(f"  Page {page}/{total_pages}: {len(entries)} entries")
            time.sleep(0.4)
        except Exception as e:
            print(f"  Page {page} error: {e}")

    total = len(all_entries)
    pending = [e for e in all_entries if e["pdf_url"] not in done_urls]
    print(f"\nTotal entries: {total}  |  Already done: {len(done_urls)}  |  To process: {len(pending)}\n")

    for i, entry in enumerate(pending, 1):
        pdf_url = entry["pdf_url"]
        title = entry["title"]
        date_listed = entry["date_listed"]

        raw_name = pdf_url.split("/")[-1]
        safe_name = re.sub(r'[<>:"/\\|?*]', "_", raw_name)
        dest = DOWNLOAD_DIR / safe_name

        print(f"[{i}/{len(pending)}] {safe_name}")
        print(f"  Title : {title[:70]}")

        # Download
        print("  Downloading...", end=" ", flush=True)
        ok = download_pdf(pdf_url, dest)
        if not ok:
            print("  Skipping (download failed).")
            continue
        print(f"ok ({dest.stat().st_size // 1024} KB)")

        # Convert to images
        print("  Converting pages to images...", end=" ", flush=True)
        images = pdf_pages_to_base64_images(dest)
        if not images:
            print("failed — no images extracted.")
            continue
        print(f"{len(images)} page(s)")

        # Analyse with Claude Vision
        print("  Sending to Claude Vision for OCR + analysis...", end=" ", flush=True)
        fields = extract_fields_with_claude_vision(client, images, title, date_listed)
        # Fallback date
        if not fields.get("date_of_judgement") or fields["date_of_judgement"] in ("Not found", "Error", ""):
            fields["date_of_judgement"] = date_listed
        print("done")

        # Write to Excel
        sr += 1
        append_row(ws, sr, fields, safe_name, pdf_url)
        wb.save(EXCEL_FILE)

        # Save progress
        done_urls.add(pdf_url)
        progress["done_urls"] = list(done_urls)
        progress["last_sr"] = sr
        save_progress(progress)

        print(f"  Case   : {fields.get('case_name','?')[:65]}")
        print(f"  Author : {fields.get('author','?')[:65]}")
        print(f"  Relief : {fields.get('relief_granted','?')[:80]}")
        print()

        time.sleep(0.5)  # be polite to the server

    wb.save(EXCEL_FILE)
    print("=" * 65)
    print(f"Complete! {sr} entries written to: {EXCEL_FILE.resolve()}")
    print(f"PDFs saved in:                    {DOWNLOAD_DIR.resolve()}")
    print("=" * 65)


if __name__ == "__main__":
    main()
